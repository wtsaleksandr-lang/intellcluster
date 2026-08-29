from __future__ import annotations

import asyncio
import re
from collections.abc import AsyncIterator
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

import httpx
from openpyxl import load_workbook

from intelligence.models import SourceRecord
from intelligence.sources.base import SourceAdapter


class CanadianImportersAdapter(SourceAdapter):
    key = "canadian_importers"
    display_name = "Canadian Importers Database"
    license_name = "Open Government Licence - Canada"
    attribution = "Innovation, Science and Economic Development Canada"

    DATASETS = {
        "major_importers_hs10": "https://ised-isde.canada.ca/site/ised/sites/default/files/documents/cid-bdic-majorimportersbyhs102023.xlsx",
        "major_importers_hs6": "https://ised-isde.canada.ca/site/ised/sites/default/files/documents/cid-bdic-majorimportersbyhs62023.xlsx",
        "major_importers_hs6_country": "https://ised-isde.canada.ca/site/ised/sites/default/files/documents/cid-bdic-majorimportersbyhs6bycountry2023.xlsx",
        "major_importers_city": "https://ised-isde.canada.ca/site/ised/sites/default/files/documents/cid-bdic-majorimportersbycity2023.xlsx",
        "major_importers_country": "https://ised-isde.canada.ca/site/ised/sites/default/files/documents/cid-bdic-majorimportersbycountry2023.xlsx",
        "hs10_descriptions": "https://ised-isde.canada.ca/site/ised/sites/default/files/documents/cid-bdic-hs10description2023.xlsx",
        "hs6_descriptions": "https://ised-isde.canada.ca/site/ised/sites/default/files/documents/cid-bdic-hs6description2023_0.xlsx",
    }

    REQUEST_HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131 Safari/537.36",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*;q=0.8",
        "Accept-Language": "en-CA,en;q=0.9",
        "Referer": "https://open.canada.ca/",
    }

    async def fetch(self, cache_dir: Path) -> list[Path]:
        target_dir = cache_dir / self.key
        target_dir.mkdir(parents=True, exist_ok=True)
        paths: list[Path] = []
        failed: list[str] = []
        timeout = httpx.Timeout(180.0, connect=30.0)

        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True, headers=self.REQUEST_HEADERS) as client:
            for name, url in self.DATASETS.items():
                path = target_dir / f"{name}.xlsx"
                if path.exists():
                    try:
                        cached = path.read_bytes()
                    except OSError:
                        cached = b""
                    if self._is_valid_xlsx(cached):
                        paths.append(path)
                        continue

                payload: bytes | None = None
                last_problem = "unknown response"
                for attempt in range(3):
                    try:
                        response = await client.get(url)
                        response.raise_for_status()
                    except httpx.HTTPError as exc:
                        last_problem = f"request failed: {exc}"
                        await asyncio.sleep(1.5 * (attempt + 1))
                        continue
                    candidate = response.content
                    content_type = response.headers.get("content-type", "")
                    if self._is_valid_xlsx(candidate):
                        payload = candidate
                        break
                    snippet = candidate[:300].decode("utf-8", errors="replace").replace("\n", " ")
                    last_problem = (
                        f"HTTP {response.status_code}, content-type={content_type!r}, bytes={len(candidate)}, "
                        f"first bytes={snippet[:180]!r}"
                    )
                    await asyncio.sleep(1.5 * (attempt + 1))

                if payload is None:
                    failed.append(name)
                    print(f"Skipping unavailable ISED workbook {name} ({last_problem})")
                    continue
                path.write_bytes(payload)
                paths.append(path)

        if not paths:
            raise RuntimeError("No Canadian Importers workbooks could be downloaded.")
        if failed:
            print(
                "Canadian Importers sync is continuing with "
                f"{len(paths)} usable workbook(s); {len(failed)} workbook(s) were unavailable."
            )
        return paths

    async def iter_records(self, paths: list[Path]) -> AsyncIterator[SourceRecord]:
        hs6_descriptions, hs10_descriptions = self._load_description_maps(paths)

        for path in paths:
            if "descriptions" in path.stem:
                continue
            if not self._is_valid_xlsx(path.read_bytes()):
                print(f"Skipping invalid cached importer workbook: {path}")
                continue

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    header_map: dict[str, int] | None = None
                    context_hs10: str | None = self._digits_from_text(sheet.title, 10)
                    context_hs6: str | None = self._digits_from_text(sheet.title, 6)
                    context_origin: str | None = None
                    blank_rows = 0

                    for values in sheet.iter_rows(values_only=True):
                        strings = [self._string(value) for value in values]
                        nonempty = [value for value in strings if value]
                        if not nonempty:
                            blank_rows += 1
                            if blank_rows >= 3:
                                header_map = None
                            continue
                        blank_rows = 0

                        row_text = " | ".join(nonempty)
                        found10 = self._digits_from_text(row_text, 10)
                        found6 = self._digits_from_text(row_text, 6)
                        if found10:
                            context_hs10 = found10
                            context_hs6 = found10[:6]
                        elif found6:
                            context_hs6 = found6

                        detected_header = self._detect_header(strings)
                        if detected_header:
                            header_map = detected_header
                            continue

                        if header_map is None:
                            if len(nonempty) == 1 and 2 <= len(nonempty[0]) <= 80:
                                candidate = nonempty[0].strip()
                                if not any(ch.isdigit() for ch in candidate) and "import" not in candidate.casefold():
                                    context_origin = candidate
                            continue

                        name = self._value_at(strings, header_map.get("name"))
                        if not name or self._looks_like_header(name):
                            continue

                        city = self._value_at(strings, header_map.get("city"))
                        province = self._value_at(strings, header_map.get("province"))
                        postal = self._value_at(strings, header_map.get("postal_code"))
                        hs10 = self._digits(self._value_at(strings, header_map.get("hs10")), 10) or context_hs10
                        hs6 = self._digits(self._value_at(strings, header_map.get("hs6")), 6) or context_hs6
                        if hs10 and not hs6:
                            hs6 = hs10[:6]
                        origin = self._value_at(strings, header_map.get("origin_country")) or context_origin
                        description = hs10_descriptions.get(hs10 or "") or hs6_descriptions.get(hs6 or "")

                        raw = {f"column_{i + 1}": value for i, value in enumerate(strings) if value}
                        record_id = "|".join(
                            part for part in [path.stem, name, hs10 or hs6 or "", origin or "", city or "", postal or ""] if part
                        )
                        yield SourceRecord(
                            source=self.key,
                            source_record_id=record_id,
                            entity_type="company",
                            name=name,
                            country="CA",
                            region=province,
                            city=city,
                            postal_code=postal,
                            source_url="https://open.canada.ca/data/en/dataset/873cfcb0-1c9b-4a48-a366-076697069bb9",
                            attributes={
                                "activity_year": 2023,
                                "hs10": hs10,
                                "hs6": hs6,
                                "product_description": description,
                                "origin_country": origin,
                                "dataset": path.stem,
                                "sheet": sheet.title,
                                "raw": raw,
                            },
                        )
            finally:
                workbook.close()

    def _load_description_maps(self, paths: list[Path]) -> tuple[dict[str, str], dict[str, str]]:
        hs6: dict[str, str] = {}
        hs10: dict[str, str] = {}
        for path in paths:
            if "descriptions" not in path.stem or not self._is_valid_xlsx(path.read_bytes()):
                continue
            target = hs10 if "hs10" in path.stem else hs6
            length = 10 if target is hs10 else 6
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    for values in sheet.iter_rows(values_only=True):
                        strings = [self._string(value) for value in values]
                        code = None
                        for value in strings:
                            code = self._digits_from_text(value, length) or self._digits(value, length)
                            if code:
                                break
                        if not code:
                            continue
                        candidates = [
                            value.strip() for value in strings
                            if value and value.strip() != code and not value.strip().isdigit() and len(value.strip()) > 3
                        ]
                        if candidates:
                            description = max(candidates, key=len)
                            if "description" not in description.casefold() and "hs code" not in description.casefold():
                                target.setdefault(code, description)
            finally:
                workbook.close()
        return hs6, hs10

    @classmethod
    def _detect_header(cls, values: list[str]) -> dict[str, int] | None:
        mapping: dict[str, int] = {}
        for index, value in enumerate(values):
            normalized = cls._norm(value)
            if not normalized:
                continue
            if normalized.startswith("companyname") or normalized.startswith("importername") or normalized in {"company", "importer", "nameofimporter"}:
                mapping["name"] = index
            elif normalized.startswith("city") or normalized == "importercity":
                mapping["city"] = index
            elif normalized.startswith("province") or normalized in {"prov", "provinceterritory"}:
                mapping["province"] = index
            elif normalized.startswith("postalcode") or normalized in {"postcode", "zipcode"}:
                mapping["postal_code"] = index
            elif normalized.startswith("countryoforigin") or normalized in {"origincountry", "country"}:
                mapping["origin_country"] = index
            elif "hs10" in normalized or normalized in {"hscode10", "tariffitem"}:
                mapping["hs10"] = index
            elif "hs6" in normalized or normalized in {"hscode6", "subheading"}:
                mapping["hs6"] = index
        return mapping if "name" in mapping else None

    @staticmethod
    def _value_at(values: list[str], index: int | None) -> str | None:
        if index is None or index < 0 or index >= len(values):
            return None
        return values[index] or None

    @classmethod
    def _looks_like_header(cls, value: str) -> bool:
        normalized = cls._norm(value)
        return normalized.startswith("companyname") or normalized.startswith("importername")

    @staticmethod
    def _is_valid_xlsx(payload: bytes) -> bool:
        if len(payload) < 4 or not payload.startswith(b"PK"):
            return False
        try:
            from io import BytesIO
            with ZipFile(BytesIO(payload)) as archive:
                names = set(archive.namelist())
                return "[Content_Types].xml" in names and "xl/workbook.xml" in names
        except BadZipFile:
            return False

    @staticmethod
    def _norm(value: Any) -> str:
        return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())

    @staticmethod
    def _string(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _digits(value: str | None, length: int) -> str | None:
        if not value:
            return None
        digits = "".join(ch for ch in value if ch.isdigit())
        if len(digits) < length:
            return None
        return digits[:length]

    @staticmethod
    def _digits_from_text(value: str | None, length: int) -> str | None:
        if not value:
            return None
        match = re.search(rf"(?<!\d)(\d{{{length}}})(?!\d)", value.replace(" ", ""))
        return match.group(1) if match else None
